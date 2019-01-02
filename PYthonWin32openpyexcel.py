from tkthread import tk, TkThread
from tkinter.filedialog import *
from tkinter import *
import tkinter.messagebox
import win32com
import win32com.client
import pandas as pd
import xlsxwriter
import os
import glob
import openpyxl
from openpyxl import load_workbook
from pandas import ExcelWriter
from PIL import ImageTk, Image
import threading
from tkinter import StringVar


root = Tk()
root.geometry("623x500+100+100")
scrollbar_in=Scrollbar(root)
scrollbar_out=Scrollbar(root)

checkimg = ImageTk.PhotoImage(Image.open(r"C:\Users\User\Desktop\excels\check-icon-3.png")) #체크 이미지
uncheckimg = ImageTk.PhotoImage(Image.open(r"C:\Users\User\Desktop\excels\check-icon-2.png")) #체크 이미지
CheckVar1=IntVar()
CheckVar2=IntVar()
state_text = StringVar()
input_filename = None
output_file = None
state_text.set("준비중")

def openinputfile():
    global input_filename
    input_filename1 = askopenfilenames(filetypes=(
        ("Excel files", "*.xls;*.xlsx"),
        ("All files", "*.*")), title="Select files")

    #print(input_filename , "\tChoice\n")
    input_excellist.delete('1.0', END)                        #텍스트 내용 삭제
    input_filename = list(input_filename1)
    if input_filename:
        CheckVar1.set("1")
    else:
        CheckVar1.set("0")

    for inputfile in input_filename:
        input_excellist.insert(INSERT, os.path.basename(inputfile)+"\n")
        #input_excellist.insert(INSERT, "\n")
    filecount = len(input_filename1)
    input_excellist.insert(INSERT, "\n\n\t\t\t")
    input_excellist.insert(INSERT, filecount)
    input_excellist.insert(INSERT, "개의 파일이 선택되었습니다.\n")
    state_text.set("파일 준비완료")

def openoutputdirectory():
    global output_file
    output_file = askopenfilename(filetypes=(
        ("XLSM files", "*.xlsm"),
        ("Excel files", "*.xls;*.xlsx"),
        ("All files", "*.*")))
    #print (output_file, "Choice\n")
    if output_file:
        CheckVar2.set("1")
        Excellist.delete('1.0', END)
        Excellist.insert(INSERT, os.path.basename(output_file)+"\n")
        #imglabel = Label(image=checkimg).grid(row=2, column=2, padx=2)
        state_text.set(os.path.basename(output_file))
    else:
        CheckVar2.set("0")
        Excellist.delete('1.0', END)
        state_text.set("준비중")

def convertxls1():
    output_browse.bind("<Button-1>", lambda e: "break")
    input_browse.bind("<Button-1>", lambda e: "break")
    convert_button.bind("<Button-1>", lambda e: "break")
    #작업중 버튼 클릭 안되게
    #print (input_filename)
    state_text.set("변환중 ...")
    Excellist.delete('1.0', END)                        #텍스트 내용 삭제
    filepath_index = "경로 ▶ "

    for f in input_filename:
        writer = pd.ExcelWriter(output_file, engine = 'openpyxl')   #출력파일 로드
        writer.book = load_workbook(output_file,keep_vba=True)      #출력파일 로드
        sheetlist = writer.book.sheetnames

        if "Sheet1" in writer.book.sheetnames:
            del writer.book["Sheet1"]
#시트1 삭제->시트1에 내용 작성->저장->매크로 호출->매크로 돌리고, 다음 루프
        getfilename = os.path.basename(f)               #파일이름 가져오기
        splitextension = getfilename.split(".")[0]      #확장자 분리
        pay_company = splitextension.split("_")[-1]     #회사이름
        pay_date = splitextension.split("_")[-2]        #날짜
        everyweek = splitextension.split("_")[-3]        #매월
        pay_s_no = splitextension.split("_")[-4]        #사업자번호
        pay_company_text = pd.DataFrame(columns=[pay_company])
        pay_date_text = pd.DataFrame(columns=[pay_date])
        everyweek_text = pd.DataFrame(columns=[everyweek])
        pay_s_no_text = pd.DataFrame(columns=[pay_s_no])
        fname = pd.DataFrame(columns=[getfilename])     #파일이름 정의
        filepath_c = pd.DataFrame(columns=[f])
        filepath = pd.DataFrame(columns=[filepath_index+f]) #"경로 ▶ " + 파일경로 정의
        df = pd.read_excel(f)       #저장할파일 읽기
        xls = pd.ExcelFile(f)

        sheetname = xls.sheet_names #시트이름 불러오기, #시트이름 리스트의 1번째([0]으로 표시)
        sheet_name_onelist = sheetname[0].split("_")[-2]    # [-1] 언더바 뒷자리
        sheet_name_twolist = sheetname[0].split("_")[-1]    # [-2] 언더바 뒷자리
        sheetn_name_one = pd.DataFrame(columns=[sheet_name_onelist])
        sheetn_name_two = pd.DataFrame(columns=[sheet_name_twolist])

        pay_s_no_text.to_excel(writer, sheet_name='Sheet1',index=False, startcol=4, startrow=4)
        pay_date_text.to_excel(writer, sheet_name='Sheet1',index=False, startcol=6, startrow=4)
        everyweek_text.to_excel(writer, sheet_name='Sheet1',index=False, startcol=5, startrow=4)
        pay_company_text.to_excel(writer, sheet_name='Sheet1',index=False, startcol=7, startrow=4)
        #fname.to_excel(writer, sheet_name='Sheet1',index=False, startcol=4, startrow=4)
        filepath.to_excel(writer, sheet_name='Sheet1',index=False, startcol=0, startrow=4)
        df.to_excel(writer, sheet_name='Sheet1', index=False, startcol=0, startrow=5) #index=False 왼쪽줄번호 제거
        sheetn_name_one.to_excel(writer, sheet_name='Sheet1',index=False, startcol=10, startrow=4)
        sheetn_name_two.to_excel(writer, sheet_name='Sheet1',index=False, startcol=11, startrow=4)

        writer.save()
        writer.close()

        xl=win32com.client.Dispatch("Excel.Application")
        xl.Workbooks.Open(Filename=output_file)
        xl.Application.Run("filecalling_new")
        xl.Workbooks(1).Close(SaveChanges=1)
        xl.Application.Quit()
        xl=0

        Excellist.insert(INSERT, os.path.basename(f))
        Excellist.insert(INSERT, "\n")
        print("Converting...\t"+f)

    print("Output : "+output_file+"\t...Save OK")
    Excellist.insert(INSERT, "\n\n\t\t\t    저장 완료\n")
    #작업끝나고 버튼 클릭 가능
    input_browse.bind("<Button-1>", lambda e: "continue")
    output_browse.bind("<Button-1>", lambda e: "continue")
    convert_button.bind("<Button-1>", lambda e: "continue")
    state_text.set("완료")

def convertxls():
    if (output_file) and (input_filename):
        t = threading.Thread(target=convertxls1)
        t.start()
    else:
        tkinter.messagebox.showwarning("오류!", " 변환파일과 저장파일을 선택해주세요.")



lbl_name = Label(root, text="변환 파일")      #파일선택라벨, 버튼
lbl_name.place(x=70, y=10)
input_browse = Button(text="Browse", command=openinputfile)
input_browse.place(x=370, y=10, width=100)
Check_infile=Checkbutton(root,text="Check",variable=CheckVar1)
Check_infile.place(x=470, y=11)
Check_infile.bind("<Button-1>", lambda e: "break")

lbl_pass = Label(root, text="저장 파일")            #파일출력라벨, 버튼
lbl_pass.place(x=70, y=40)
output_browse = Button(text="Browse", command=openoutputdirectory)
output_browse.place(x=370, y=40, width=100)
Check_outfile=Checkbutton(root,text="Check",variable=CheckVar2)
Check_outfile.place(x=470, y=41)
Check_outfile.bind("<Button-1>", lambda e: "break")

lbl_state = Label(root, text="State  : ")            #state 라벨
lbl_state.place(x=70, y=70)

state_lbl = Label(root, text="state_text", textvariable = state_text)            #state 라벨
state_lbl.place(x=170, y=70)


input_excellist = Text(root, height=13, width=68, relief=SOLID, yscrollcommand=scrollbar_in.set)     #불러온 엑셀파일 텍스트박스1에 출력
input_excellist.place(x=25, y=95)
input_excellist.bind("<Key>", lambda e: "break")      #키 입력 안되게
input_excellist.bind("<Button-1>", lambda e: "break")      #마우스 입력 안되게
scrollbar_in.config(command=input_excellist.yview)     #스크롤바
scrollbar_in.place(x=505, y=95, height=175)

Excellist = Text(root, height=14, width=68, relief=SOLID, yscrollcommand=scrollbar_out.set)     #변환한 엑셀파일 텍스트박스에 출력
Excellist.place(x=25, y=285)
Excellist.bind("<Key>", lambda e: "break")      #키 입력 안되게
Excellist.bind("<Button-1>", lambda e: "break")      #마우스 입력 안되게
scrollbar_out.config(command=input_excellist.yview)     #스크롤바
scrollbar_out.place(x=505, y=295, height=175)

convert_button = Button(text="변환하기", bg="#FF9966", fg="black", command=convertxls)
convert_button.place(x=530, y=180, width=75, height=180)

root.title("일용 노무비")
root.iconbitmap(r"C:\Users\User\Desktop\123.ico")
root.resizable(False, False)
root.mainloop()
